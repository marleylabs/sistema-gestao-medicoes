from __future__ import annotations

import argparse
import base64
import hashlib
import json
import os
import unicodedata
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import pandas as pd
from cryptography.hazmat.primitives.ciphers.aead import AESGCM
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
from sqlalchemy import MetaData, bindparam, create_engine, text
from sqlalchemy.dialects.postgresql import insert


DEFAULT_SHEET = "Documentos"
DEFAULT_BASE_SHEET = "Base"
DEFAULT_PAYMENT_MAP_SHEET = "MAPA PAGTO"
DEFAULT_BM_AUX_SHEET = "Documentos Auxiliares"
DEFAULT_EXCEL_PATH = r"C:\Users\anderson.marley\Desktop\2605 - ELABORAÇÃO DE BMS - ( PROJETISTAS ).xlsm"
SHEET_ALIASES = {
    "Documentos": ["Documentos", "Geral"],
    "Base": ["Base", "BASE"],
    "MAPA PAGTO": ["MAPA PAGTO"],
    "Documentos Auxiliares": ["Documentos Auxiliares", "BM AUX", "BM - AUX"],
}
BM_AUX_ALLOWED_COLLABORATORS = {
    "mauricio spindola",
    "cristiano jeferson",
}
ENCRYPTED_PREFIX = "enc:v1"
SENSITIVE_RAW_COLUMNS = {
    "cpf",
    "cnpj",
    "cpf/cnpj",
    "cpf / cnpj",
    "e-mail",
    "email",
}

PROJECT_COLUMNS = {
    "codigo_projeto": ["Projeto Referente"],
    "titulo_primario": ["Título Primário"],
    "centro_custo": ["Centro de Custo"],
    "localizacao": ["Localização", "Localização.1", "Localização .1"],
    "contrato": ["CONTRATO"],
}

PROFESSIONAL_COLUMNS = {
    "nome": ["PROJETISTA", "Evidência", "Número do Documento"],
    "funcao": ["FUNÇÃO", "FUNÇÃO2", "Função"],
}

COORDINATOR_COLUMNS = {
    "nome": ["Coordenador", "Mesclado"],
}

BASE_PROFESSIONAL_COLUMNS = {
    "codigo": ["Codigo", "Código"],
    "nome_completo": ["Nome Completo"],
    "cpf": ["CPF"],
    "razao_social": ["Razão Social"],
    "cnpj": ["CNPJ"],
    "email": ["E-mail", "Email"],
    "funcao": ["Função", "FUNÇÃO"],
}

PAYMENT_MAP_COLUMNS = {
    "status_source": ["ATO"],
    "codigo": ["PROJETISTA"],
    "valor": ["VALOR"],
}

PAYMENT_MAP_ITEM_COLUMNS = {
    "ato": ["ATO"],
    "projetista_codigo": ["PROJETISTA"],
    "responsavel": ["RESPONSÁVEL", "RESPONSAVEL"],
    "cpf_cnpj": ["CPF / CNPJ", "CPF/CNPJ"],
    "razao_social": ["CNPJ", "Razão Social", "RAZÃO SOCIAL"],
    "intr_sossego": ["Intr. Sossego", "Intr. Sossego."],
    "salobo": ["Salobo"],
    "acg": ["ACG"],
    "escadas_alumar": ["Escadas Alumar"],
    "valor": ["VALOR"],
    "rev": ["REV"],
    "status": ["STATUS", "CONDIÇÃO"],
}

MEASUREMENT_COLUMNS = {
    "numero_medicao": ["Número da Medição"],
    "mesclado": ["Mesclado"],
    "numero_documento": ["Número do Documento"],
    "evidencia": ["Evidência"],
    "data_cadastro": ["Data de Cadastro"],
    "formato": ["Formato", "Coluna1"],
    "quantidade": ["Quantidade"],
    "multiplicador": ["Multiplicador"],
    "equivalente_a1_horas": ["Equivalente (A1 ou Horas)"],
    "porcentagem_revisao": ["Porcentagem de Revisão"],
    "emissao_inicial": ["Emissão Inicial"],
    "retorno_vale": ["Retorno Vale"],
    "encerramento": ["Encerramento"],
    "arquivamento": ["Arquivamento"],
    "medido_horas": ["Medido (Horas)"],
    "item_qqp": ["Item da QQP"],
    "valor_unitario": ["Valor Unitário"],
    "valor_bruto": ["Valor Bruto"],
    "valor_total": ["Valor Total"],
    "obs": ["OBS"],
    "motivo_desconto": [
        "Motivo Desconto",
        "MOTIVO DESCONTO",
        "Descrição do Desconto",
        "DESCRIÇÃO DO DESCONTO",
        "Descricao do Desconto",
        "DESCRICAO DO DESCONTO",
    ],
    "valor_desconto": [
        "Valor Desconto",
        "VALOR DESCONTO",
        "Valor do Desconto",
        "VALOR DO DESCONTO",
    ],
    "valor_reajuste": ["Valor do Reajuste"],
    "ciclo": ["CICLO"],
    "referencia": ["REFERÊNCIA"],
    "percentual_emissao": ["% EMISSÃO"],
    "tipo2": ["TIPO", "TIPO2"],
    "condicao": ["CONDIÇÃO"],
    "valor_medicao": ["VALOR DE MEDIÇÃO"],
}

BM_AUX_COLUMNS = {
    "projeto": ["Projeto"],
    "fase": ["Fase"],
    "numero_cliente": ["Num_Cliente", "Número Cliente", "Numero Cliente"],
    "responsavel": ["Responsavel", "Responsável"],
    "auxiliar": ["Auxiliar"],
    "data_entrega": ["Data_Entrega", "Data Entrega"],
    "tipo_revisao": ["Tipo_Revisao", "Tipo Revisão"],
    "tipo_emissao": ["Tipo_Emissao", "Tipo Emissão"],
    "data_emissao": ["Data_Emissao", "Data Emissão"],
    "evidencia_emissao": ["Evidencia_Emissao", "Evidência Emissão"],
    "status_retorno": ["Status_Retorno", "Status Retorno"],
    "formato": ["Formato"],
    "quantidade": ["Quantidade"],
    "perc_revisao": ["Perc_Revisao", "Perc Revisão"],
    "equivalente_revisado": ["Equivalente_Revisado", "Equivalente Revisado"],
    "tipo_doc": ["Tipo_Doc", "Tipo Doc", "TIPO DG/DOC/HH"],
    "contrato": ["Contrato"],
    "orcamento": ["Orçamento", "Orcamento"],
    "ordem_emissao": ["Ordem_Emissao", "Ordem Emissão"],
    "ultima_emissao": ["Ultima_Emissao", "Última_Emissao", "Última Emissão"],
    "ciclo": ["Ciclo", "CICLO"],
    "ciclo_retorno": ["CICLO RETORNO", "Ciclo Retorno"],
    "mesclado": ["Mesclado"],
    "valor": ["VALOR"],
    "percentual_emissao": ["% EMISSÃO", "% EMISSAO"],
    "valor_medicao": ["VALOR DA MEDIÇÃO", "VALOR DA MEDICAO", "Valor De Medição"],
}


def is_valid_measurement_key(numero_medicao: str | None, codigo_projeto: str | None) -> bool:
    if not numero_medicao or not codigo_projeto:
        return False

    invalid_values = {
        "número da medição",
        "numero da medição",
        "projeto referente",
        "coordenador",
        "projetista",
    }
    if numero_medicao.casefold() in invalid_values or codigo_projeto.casefold() in invalid_values:
        return False

    return numero_medicao.upper().startswith("BM")


def first_value(row: pd.Series, candidates: list[str]) -> Any:
    for column in candidates:
        if column not in row.index:
            continue

        value = row[column]
        if isinstance(value, pd.Series):
            value = next((item for item in value.tolist() if not pd.isna(item) and not (isinstance(item, str) and not item.strip())), None)

        if not pd.isna(value):
            if isinstance(value, str) and not value.strip():
                continue
            return value

    normalized_candidates = {normalize_for_compare(column) for column in candidates}
    for column in row.index:
        if normalize_for_compare(str(column)) not in normalized_candidates:
            continue

        value = row[column]
        if isinstance(value, pd.Series):
            value = next((item for item in value.tolist() if not pd.isna(item) and not (isinstance(item, str) and not item.strip())), None)

        if not pd.isna(value):
            if isinstance(value, str) and not value.strip():
                continue
            return value
    return None


def clean_text(value: Any) -> str | None:
    if value is None or pd.isna(value):
        return None
    value = str(value).replace("\xa0", " ").strip()
    value = " ".join(value.split())
    return value or None


def encryption_key() -> bytes:
    encoded = os.getenv("DATA_ENCRYPTION_KEY")
    if not encoded:
        raise RuntimeError("DATA_ENCRYPTION_KEY não configurada.")
    key = base64.b64decode(encoded)
    if len(key) != 32:
        raise RuntimeError("DATA_ENCRYPTION_KEY deve conter exatamente 32 bytes em Base64.")
    return key


def encrypt_sensitive(value: Any) -> str | None:
    cleaned = clean_text(value)
    if not cleaned:
        return None
    nonce = os.urandom(12)
    encrypted = AESGCM(encryption_key()).encrypt(nonce, cleaned.encode("utf-8"), None)
    ciphertext, tag = encrypted[:-16], encrypted[-16:]
    return ":".join(
        [
            ENCRYPTED_PREFIX,
            base64.b64encode(nonce).decode("ascii"),
            base64.b64encode(tag).decode("ascii"),
            base64.b64encode(ciphertext).decode("ascii"),
        ]
    )


def safe_raw_payload(row: pd.Series, unnamed_prefix: str) -> dict[str, Any]:
    return {
        str(key): json_safe(value)
        for key, value in row.to_dict().items()
        if not str(key).startswith(unnamed_prefix)
        and normalize_for_compare(str(key)) not in SENSITIVE_RAW_COLUMNS
    }


def normalize_for_compare(value: str | None) -> str:
    if value is None:
        return ""
    normalized = unicodedata.normalize("NFKD", value)
    normalized = "".join(char for char in normalized if not unicodedata.combining(char))
    normalized = normalized.replace("\xa0", " ")
    return " ".join(normalized.casefold().strip().split())


def should_keep_vba(excel_path: Path) -> bool:
    return excel_path.suffix.casefold() == ".xlsm"


def resolve_sheet_name(excel_path: Path, requested_name: str, aliases: list[str] | None = None) -> str:
    workbook = load_workbook(excel_path, read_only=False, data_only=True, keep_vba=should_keep_vba(excel_path))
    try:
        available = {normalize_for_compare(name): name for name in workbook.sheetnames}
    finally:
        workbook.close()

    candidates = [requested_name, *(aliases or [])]
    seen: set[str] = set()
    for candidate in candidates:
        normalized = normalize_for_compare(candidate)
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)
        if normalized in available:
            return available[normalized]

    raise ValueError(
        f"Aba '{requested_name}' não encontrada. Abas disponíveis: {', '.join(available.values())}."
    )


def resolve_optional_sheet_name(excel_path: Path, requested_name: str, aliases: list[str] | None = None) -> str | None:
    try:
        return resolve_sheet_name(excel_path, requested_name, aliases)
    except ValueError:
        return None


def normalize_collaborator_status(value: Any) -> str:
    source = clean_text(value)
    if normalize_for_compare(source) == "producao":
        return "PRODUÇÃO"
    return "ATO"


def is_expense_professional_name(value: str | None) -> bool:
    return normalize_for_compare(value).startswith("despesa alimentacao")


def has_positive_payment_value(valor: Decimal | None) -> bool:
    return bool(valor is not None and valor > 0)


def is_payment_map_ato(value: Any, valor: Decimal | None, codigo: str | None, positive_payment_codes: set[str]) -> bool:
    normalized = normalize_for_compare(clean_text(value))
    return bool(normalized and normalized != "producao" and has_positive_payment_value(valor))


def is_payment_map_production(value: Any, valor: Decimal | None, codigo: str | None, positive_payment_codes: set[str]) -> bool:
    normalized = normalize_for_compare(clean_text(value))
    has_positive_source_value = normalize_for_compare(codigo) in positive_payment_codes
    return bool(normalized in {"", "producao"} and (has_positive_payment_value(valor) or has_positive_source_value))


def clean_decimal(value: Any, default: Decimal | None = Decimal("0")) -> Decimal | None:
    if value is None or pd.isna(value):
        return default
    if isinstance(value, (int, float, Decimal)):
        if pd.isna(value):
            return default
        return Decimal(str(value))

    raw = str(value).strip()
    if not raw:
        return default

    normalized = raw.replace("R$", "").replace("\xa0", "").replace(" ", "")
    if normalized.endswith("%"):
        normalized = normalized[:-1]

    if "," in normalized and "." in normalized:
        normalized = normalized.replace(".", "").replace(",", ".")
    elif "," in normalized:
        normalized = normalized.replace(",", ".")

    try:
        return Decimal(normalized)
    except InvalidOperation:
        return default


def clean_percent(value: Any) -> Decimal | None:
    number = clean_decimal(value, default=None)
    if number is None:
        return None
    if isinstance(value, str) and "%" in value:
        return number / Decimal("100")
    return number


def clean_date(value: Any) -> date | None:
    if value is None or pd.isna(value):
        return None
    parsed = pd.to_datetime(value, dayfirst=True, errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed.date()


def only_digits(value: Any) -> str:
    if value is None or pd.isna(value):
        return ""
    return "".join(char for char in str(value) if char.isdigit())


def format_cnpj(value: Any) -> str | None:
    digits = only_digits(value)
    if len(digits) == 13:
        digits = digits.zfill(14)
    if len(digits) != 14:
        return clean_text(value)
    return f"{digits[:2]}.{digits[2:5]}.{digits[5:8]}/{digits[8:12]}-{digits[12:]}"


def json_safe(value: Any) -> Any:
    if value is None or pd.isna(value):
        return None
    if isinstance(value, pd.Timestamp):
        return value.isoformat()
    if isinstance(value, (date, Decimal)):
        return str(value)
    return value


def source_hash(payload: dict[str, Any]) -> str:
    encoded = json.dumps(payload, ensure_ascii=False, sort_keys=True, default=str)
    return hashlib.sha256(encoded.encode("utf-8")).hexdigest()


def extract(row: pd.Series, mapping: dict[str, list[str]]) -> dict[str, Any]:
    return {target: first_value(row, candidates) for target, candidates in mapping.items()}


def load_schema(engine, schema_path: Path) -> None:
    sql = schema_path.read_text(encoding="utf-8")
    with engine.begin() as conn:
        conn.execute(text(sql))


def reflect_tables(engine):
    metadata = MetaData()
    metadata.reflect(
        engine,
        only=[
            "projetos",
            "profissionais",
            "medicoes",
            "mapa_pagamento_contexto",
            "mapa_pagamento_itens",
            "bm_aux_medicoes",
        ],
    )
    return (
        metadata.tables["projetos"],
        metadata.tables["profissionais"],
        metadata.tables["medicoes"],
        metadata.tables["mapa_pagamento_contexto"],
        metadata.tables["mapa_pagamento_itens"],
        metadata.tables["bm_aux_medicoes"],
    )


def clear_imported_database(conn, ciclo: str) -> None:
    # Substitui somente o ciclo carregado pela planilha.
    # Os demais ciclos, cadastros compartilhados, usuários internos e contratos
    # manuais são preservados para permitir histórico multi-ciclo.
    params = {"ciclo": ciclo}
    conn.execute(
        text(
            """
            delete from sgc_logs
            where ciclo = :ciclo
               or sgc_id in (
                    select id
                    from sgc_aprovacoes_medicao
                    where ciclo = :ciclo
               )
            """
        ),
        params,
    )
    conn.execute(text("delete from sgc_aprovacoes_medicao where ciclo = :ciclo"), params)
    conn.execute(text("delete from medicoes where ciclo = :ciclo"), params)
    conn.execute(text("delete from mapa_pagamento_itens where ciclo = :ciclo"), params)
    conn.execute(text("delete from mapa_pagamento_contexto where ciclo = :ciclo"), params)
    conn.execute(text("delete from bm_aux_medicoes where ciclo = :ciclo"), params)
    conn.execute(text("delete from etl_execucoes where ciclo = :ciclo"), params)


def matches_any_collaborator(value: Any, normalized_codes: set[str]) -> bool:
    normalized = normalize_for_compare(clean_text(value))
    if not normalized:
        return False
    if normalized in normalized_codes:
        return True
    return any(same_person_name(normalized, candidate) for candidate in normalized_codes)


def collect_import_collaborator_codes(
    df: pd.DataFrame,
    bm_aux_df: pd.DataFrame,
    canonical_codes: dict[str, str],
    ciclo: str,
) -> set[str]:
    codes: set[str] = set()

    def add_code(value: Any) -> None:
        cleaned = clean_text(value)
        if not cleaned:
            return
        canonical = canonical_codes.get(normalize_for_compare(cleaned), cleaned)
        if not is_expense_professional_name(canonical):
            codes.add(canonical)

    for _, row in df.iterrows():
        project_raw = extract(row, PROJECT_COLUMNS)
        codigo_projeto = clean_text(project_raw["codigo_projeto"])
        numero_medicao = clean_text(first_value(row, MEASUREMENT_COLUMNS["numero_medicao"]))
        valid_measurement = is_valid_measurement_key(numero_medicao, codigo_projeto)
        discount_only = not valid_measurement and has_discount_data(row)
        if not valid_measurement and not discount_only:
            continue

        professional_raw = extract(row, PROFESSIONAL_COLUMNS)
        if is_bm_aux_only_collaborator(professional_raw["nome"], canonical_codes) and not discount_only:
            continue
        add_code(professional_raw["nome"])

    for _, row in bm_aux_df.iterrows():
        raw = extract(row, BM_AUX_COLUMNS)
        row_cycle = normalize_cycle(raw["ciclo"])
        if row_cycle and row_cycle != ciclo:
            continue
        for _, codigo in bm_aux_people(raw, canonical_codes):
            add_code(codigo)

    return codes


def clear_imported_collaborators(conn, ciclo: str, collaborator_codes: set[str]) -> None:
    # Atualização incremental: remove somente os dados do ciclo pertencentes aos
    # fornecedores presentes na planilha atual. Fornecedores ausentes no arquivo
    # permanecem preservados no ciclo.
    normalized_codes = {normalize_for_compare(code) for code in collaborator_codes if clean_text(code)}
    if not normalized_codes:
        return

    professional_rows = conn.execute(
        text("select id, nome, codigo, nome_completo from profissionais")
    ).mappings().all()
    professional_ids = [
        str(row["id"])
        for row in professional_rows
        if (
            matches_any_collaborator(row.get("codigo"), normalized_codes)
            or matches_any_collaborator(row.get("nome"), normalized_codes)
            or matches_any_collaborator(row.get("nome_completo"), normalized_codes)
        )
    ]

    payment_rows = conn.execute(
        text(
            """
            select id, projetista_codigo, responsavel, razao_social
            from mapa_pagamento_itens
            where ciclo = :ciclo
            """
        ),
        {"ciclo": ciclo},
    ).mappings().all()
    payment_ids = [
        str(row["id"])
        for row in payment_rows
        if (
            matches_any_collaborator(row.get("projetista_codigo"), normalized_codes)
            or matches_any_collaborator(row.get("responsavel"), normalized_codes)
            or matches_any_collaborator(row.get("razao_social"), normalized_codes)
        )
    ]

    bm_aux_rows = conn.execute(
        text("select id, responsavel_codigo from bm_aux_medicoes where ciclo = :ciclo"),
        {"ciclo": ciclo},
    ).mappings().all()
    bm_aux_ids = [
        str(row["id"])
        for row in bm_aux_rows
        if matches_any_collaborator(row.get("responsavel_codigo"), normalized_codes)
    ]

    sgc_rows = conn.execute(
        text(
            """
            select id, colaborador_codigo, colaborador_nome
            from sgc_aprovacoes_medicao
            where ciclo = :ciclo
            """
        ),
        {"ciclo": ciclo},
    ).mappings().all()
    sgc_ids = [
        str(row["id"])
        for row in sgc_rows
        if (
            matches_any_collaborator(row.get("colaborador_codigo"), normalized_codes)
            or matches_any_collaborator(row.get("colaborador_nome"), normalized_codes)
        )
    ]

    if sgc_ids:
        conn.execute(
            text("delete from sgc_logs where sgc_id in :ids").bindparams(bindparam("ids", expanding=True)),
            {"ids": sgc_ids},
        )
        conn.execute(
            text("delete from sgc_aprovacoes_medicao where id in :ids").bindparams(bindparam("ids", expanding=True)),
            {"ids": sgc_ids},
        )

    if professional_ids:
        conn.execute(
            text("delete from medicoes where ciclo = :ciclo and id_profissional in :ids").bindparams(bindparam("ids", expanding=True)),
            {"ciclo": ciclo, "ids": professional_ids},
        )

    if payment_ids:
        conn.execute(
            text("delete from mapa_pagamento_itens where id in :ids").bindparams(bindparam("ids", expanding=True)),
            {"ids": payment_ids},
        )

    if bm_aux_ids:
        conn.execute(
            text("delete from bm_aux_medicoes where id in :ids").bindparams(bindparam("ids", expanding=True)),
            {"ids": bm_aux_ids},
        )


def upsert_project(conn, projetos, data: dict[str, Any]) -> int:
    stmt = insert(projetos).values(**data)
    stmt = (
        stmt
        .on_conflict_do_update(
            index_elements=[projetos.c.codigo_projeto],
            set_={
                "titulo_primario": stmt.excluded.titulo_primario,
                "centro_custo": stmt.excluded.centro_custo,
                "localizacao": stmt.excluded.localizacao,
                "contrato": stmt.excluded.contrato,
                "updated_at": text("now()"),
            },
        )
        .returning(projetos.c.id)
    )
    return conn.execute(stmt).scalar_one()


def upsert_professional(conn, profissionais, data: dict[str, Any]) -> int | None:
    nome = clean_text(data.get("nome"))
    if not nome or is_expense_professional_name(nome):
        return None

    payload = {
        "nome": nome,
        "codigo": clean_text(data.get("codigo")),
        "nome_completo": clean_text(data.get("nome_completo")),
        "cpf": encrypt_sensitive(data.get("cpf")),
        "razao_social": clean_text(data.get("razao_social")),
        "cnpj": encrypt_sensitive(data.get("cnpj")),
        "email": encrypt_sensitive(data.get("email")),
        "status_colaborador": clean_text(data.get("status_colaborador")),
        "funcao": clean_text(data.get("funcao")),
    }
    stmt = insert(profissionais).values(**payload)
    stmt = (
        stmt
        .on_conflict_do_update(
            index_elements=[profissionais.c.nome],
            set_={
                "codigo": text("coalesce(excluded.codigo, profissionais.codigo)"),
                "nome_completo": text("coalesce(excluded.nome_completo, profissionais.nome_completo)"),
                "cpf": text("coalesce(excluded.cpf, profissionais.cpf)"),
                "razao_social": text("coalesce(excluded.razao_social, profissionais.razao_social)"),
                "cnpj": text("coalesce(excluded.cnpj, profissionais.cnpj)"),
                "email": text("coalesce(excluded.email, profissionais.email)"),
                "status_colaborador": text("coalesce(excluded.status_colaborador, profissionais.status_colaborador)"),
                "funcao": text("coalesce(profissionais.funcao, excluded.funcao)"),
                "updated_at": text("now()"),
            },
        )
        .returning(profissionais.c.id)
    )
    return conn.execute(stmt).scalar_one()


def upsert_base_professional(conn, profissionais, data: dict[str, Any]) -> int:
    stmt = insert(profissionais).values(**data)
    stmt = (
        stmt
        .on_conflict_do_update(
            index_elements=[profissionais.c.nome],
            set_={
                "codigo": stmt.excluded.codigo,
                "nome_completo": stmt.excluded.nome_completo,
                "cpf": stmt.excluded.cpf,
                "razao_social": stmt.excluded.razao_social,
                "cnpj": stmt.excluded.cnpj,
                "email": stmt.excluded.email,
                "status_colaborador": text("coalesce(profissionais.status_colaborador, excluded.status_colaborador)"),
                "funcao": stmt.excluded.funcao,
                "updated_at": text("now()"),
            },
        )
        .returning(profissionais.c.id)
    )
    return conn.execute(stmt).scalar_one()


def upsert_payment_map_status(conn, profissionais, data: dict[str, Any]) -> int:
    stmt = insert(profissionais).values(**data)
    stmt = (
        stmt
        .on_conflict_do_update(
            index_elements=[profissionais.c.nome],
            set_={
                "codigo": text("coalesce(profissionais.codigo, excluded.codigo)"),
                "nome_completo": text("coalesce(excluded.nome_completo, profissionais.nome_completo)"),
                "status_colaborador": stmt.excluded.status_colaborador,
                "updated_at": text("now()"),
            },
        )
        .returning(profissionais.c.id)
    )
    return conn.execute(stmt).scalar_one()


def build_base_professional(row: pd.Series) -> dict[str, Any] | None:
    raw = extract(row, BASE_PROFESSIONAL_COLUMNS)
    codigo = clean_text(raw["codigo"])
    if not codigo or codigo.casefold() in {"codigo", "código"}:
        return None

    return {
        "nome": codigo,
        "codigo": codigo,
        "nome_completo": clean_text(raw["nome_completo"]),
        "cpf": encrypt_sensitive(raw["cpf"]),
        "razao_social": clean_text(raw["razao_social"]),
        "cnpj": encrypt_sensitive(raw["cnpj"]),
        "email": encrypt_sensitive(raw["email"]),
        "status_colaborador": None,
        "funcao": clean_text(raw["funcao"]),
    }


def build_positive_payment_codes(df: pd.DataFrame) -> set[str]:
    positive_payment_codes: set[str] = set()

    for _, row in df.iterrows():
        codigo = clean_text(first_value(row, ["PROJETISTA"]))
        valor_medicao = clean_decimal(first_value(row, MEASUREMENT_COLUMNS["valor_medicao"]))
        if codigo and has_positive_payment_value(valor_medicao):
            positive_payment_codes.add(normalize_for_compare(codigo))

    return positive_payment_codes


def build_canonical_professional_codes(base_df: pd.DataFrame) -> dict[str, str]:
    canonical_codes: dict[str, str] = {}
    for _, row in base_df.iterrows():
        raw = extract(row, BASE_PROFESSIONAL_COLUMNS)
        codigo = clean_text(raw["codigo"])
        if not codigo:
            continue
        canonical_codes[normalize_for_compare(codigo)] = codigo
        nome_completo = clean_text(raw["nome_completo"])
        if nome_completo:
            canonical_codes[normalize_for_compare(nome_completo)] = codigo
    return canonical_codes


def build_payment_map_canonical_codes(payment_map_df: pd.DataFrame) -> dict[str, str]:
    canonical_codes: dict[str, str] = {}
    for _, row in payment_map_df.iterrows():
        raw = extract(row, PAYMENT_MAP_ITEM_COLUMNS)
        codigo = clean_text(raw["projetista_codigo"])
        if not codigo:
            continue

        normalized_code = normalize_for_compare(codigo)
        if normalized_code in {"projetista", "total", "valor"}:
            continue

        canonical_codes[normalized_code] = codigo
        responsavel = clean_text(raw["responsavel"])
        if responsavel:
            canonical_codes[normalize_for_compare(responsavel)] = codigo
    return canonical_codes


def build_payment_map_status(
    row: pd.Series,
    positive_payment_codes: set[str],
    canonical_codes: dict[str, str],
) -> dict[str, Any] | None:
    raw = extract(row, PAYMENT_MAP_COLUMNS)
    item_raw = extract(row, PAYMENT_MAP_ITEM_COLUMNS)
    codigo = clean_text(raw["codigo"])
    status_source = clean_text(raw["status_source"])
    valor = clean_decimal(raw["valor"])
    if not codigo:
        return None

    invalid_codes = {"projetista", "total", "valor"}
    if normalize_for_compare(codigo) in invalid_codes:
        return None

    codigo = canonical_codes.get(normalize_for_compare(codigo), codigo)

    if is_payment_map_ato(status_source, valor, codigo, positive_payment_codes):
        status_colaborador = "ATO"
    elif is_payment_map_production(status_source, valor, codigo, positive_payment_codes):
        status_colaborador = "PRODUÇÃO"
    else:
        return None

    return {
        "nome": codigo,
        "codigo": codigo,
        "nome_completo": clean_text(item_raw["responsavel"]),
        "status_colaborador": status_colaborador,
    }


def derive_ciclo_from_mes_referencia(mes_referencia: str | None) -> str | None:
    """Deriva o ciclo YYMM a partir de strings como 'Maio de 2026' ou 'Maio/2026'.
    Retorna None se não conseguir derivar com certeza — nunca usa fallback silencioso."""
    import re
    if not mes_referencia:
        return None
    meses = {
        "janeiro": "01", "fevereiro": "02", "março": "03", "marco": "03",
        "abril": "04", "maio": "05", "junho": "06",
        "julho": "07", "agosto": "08", "setembro": "09",
        "outubro": "10", "novembro": "11", "dezembro": "12",
    }
    texto = mes_referencia.lower().strip()
    mes_num = next((v for k, v in meses.items() if k in texto), None)
    # Busca o último grupo de 4 dígitos que representa um ano razoável (2020-2099)
    anos = re.findall(r"\b(20[2-9]\d)\b", texto)
    if not mes_num or not anos:
        return None
    ano = anos[-1][-2:]  # últimos 2 dígitos do ano mais recente encontrado
    ciclo = f"{ano}{mes_num}"
    # Validação básica: YYMM onde MM entre 01-12
    if not re.fullmatch(r"\d{2}(0[1-9]|1[0-2])", ciclo):
        return None
    return ciclo


def build_payment_map_context(excel_path: Path, sheet_name: str, ciclo: str | None = None) -> dict[str, Any]:
    workbook = load_workbook(excel_path, read_only=False, data_only=True, keep_vba=should_keep_vba(excel_path))
    try:
        sheet = workbook[sheet_name]
        contract_columns = range(7, 12)
        first_row_headers = {
            normalize_for_compare(clean_text(sheet.cell(1, column).value))
            for column in range(1, sheet.max_column + 1)
            if clean_text(sheet.cell(1, column).value)
        }
        header_only_layout = {"ato", "projetista", "valor"}.issubset(first_row_headers)

        contratos = []
        rateio = []
        if not header_only_layout:
            for column in contract_columns:
                nome = clean_text(sheet.cell(4, column).value)
                if not nome:
                    continue
                contratos.append(
                    {
                        "contrato": nome,
                        "valor": float(clean_decimal(sheet.cell(5, column).value) or 0),
                    }
                )
                rateio.append(
                    {
                        "contrato": nome,
                        "percentual": float(clean_decimal(sheet.cell(6, column).value) or 0),
                    }
                )

        mes_referencia = None if header_only_layout else clean_text(sheet["H1"].value)
        if not ciclo:
            ciclo = derive_ciclo_from_mes_referencia(mes_referencia)
        if not ciclo:
            raise ValueError(
                f"Não foi possível derivar o ciclo a partir de '{mes_referencia}'. "
                "Informe o ciclo explicitamente (ex: 2606) no campo da interface."
            )

        return {
            "ciclo": ciclo,
            "mes_referencia": mes_referencia,
            "producao_label": None if header_only_layout else clean_text(sheet["G2"].value),
            "producao_inicio": None if header_only_layout else clean_date(sheet["H2"].value),
            "producao_fim": None if header_only_layout else clean_date(sheet["K2"].value),
            "ato_label": None if header_only_layout else clean_text(sheet["H3"].value),
            "ato_ciclo": None if header_only_layout else clean_text(sheet["K3"].value),
            "contratos": contratos,
            "rateio": rateio,
        }
    finally:
        workbook.close()


def build_generated_payment_context(df: pd.DataFrame, bm_aux_df: pd.DataFrame, ciclo: str | None = None) -> dict[str, Any]:
    if not ciclo:
        ciclos: list[str] = []
        for source_df, mapping in [(df, MEASUREMENT_COLUMNS), (bm_aux_df, BM_AUX_COLUMNS)]:
            if source_df.empty:
                continue
            for _, row in source_df.iterrows():
                raw = extract(row, mapping)
                row_cycle = normalize_cycle(raw.get("ciclo"))
                if row_cycle:
                    ciclos.append(row_cycle)
        ciclo = next((value for value in ciclos if value), None)
    if not ciclo:
        raise ValueError("Informe o ciclo explicitamente no campo da interface para importar sem MAPA PAGTO.")

    contratos_por_nome: dict[str, Decimal] = {}
    for source_df, mapping, value_key in [
        (df, MEASUREMENT_COLUMNS, "valor_medicao"),
        (bm_aux_df, BM_AUX_COLUMNS, "valor_medicao"),
    ]:
        if source_df.empty:
            continue
        for _, row in source_df.iterrows():
            raw = extract(row, mapping)
            contrato = clean_text(raw.get("contrato"))
            if not contrato:
                continue
            valor = clean_decimal(raw.get(value_key))
            if valor < 0:
                continue
            contratos_por_nome[contrato] = contratos_por_nome.get(contrato, Decimal("0")) + valor

    total_contratos = sum(contratos_por_nome.values(), Decimal("0"))
    contratos = [
        {"contrato": nome, "valor": float(valor)}
        for nome, valor in sorted(contratos_por_nome.items())
    ]
    rateio = [
        {
            "contrato": nome,
            "percentual": float(valor / total_contratos) if total_contratos > 0 else 0,
        }
        for nome, valor in sorted(contratos_por_nome.items())
    ]

    return {
        "ciclo": ciclo,
        "mes_referencia": None,
        "producao_label": "PRODUÇÃO",
        "producao_inicio": None,
        "producao_fim": None,
        "ato_label": "ATO",
        "ato_ciclo": ciclo,
        "contratos": contratos,
        "rateio": rateio,
    }


def read_excel_table(excel_path: Path, sheet_name: str, table_name: str | list[str]) -> pd.DataFrame:
    workbook = load_workbook(excel_path, read_only=False, data_only=True, keep_vba=should_keep_vba(excel_path))
    try:
        sheet = workbook[sheet_name]
        table_names = [table_name] if isinstance(table_name, str) else table_name
        table_key = next((name for name in table_names if name in sheet.tables), None)
        if table_key is None:
            return pd.DataFrame()

        table = sheet.tables[table_key]
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        rows = [
            list(row)
            for row in sheet.iter_rows(
                min_row=min_row,
                max_row=max_row,
                min_col=min_col,
                max_col=max_col,
                values_only=True,
            )
        ]
        return dataframe_from_excel_rows(rows)
    finally:
        workbook.close()


def dataframe_from_excel_rows(rows: list[list[Any]]) -> pd.DataFrame:
    if not rows:
        return pd.DataFrame()

    headers = [clean_text(value) or f"unnamed_{index + 1}" for index, value in enumerate(rows[0])]
    if len(rows) < 2:
        return pd.DataFrame(columns=headers)

    df = pd.DataFrame(rows[1:], columns=headers)
    return df.dropna(how="all").reset_index(drop=True)


def read_excel_header_region(
    excel_path: Path,
    sheet_name: str,
    required_headers: set[str],
    key_header: str,
    max_header_row: int = 100,
    row_key_headers: list[str] | None = None,
) -> pd.DataFrame:
    workbook = load_workbook(excel_path, read_only=False, data_only=True, keep_vba=should_keep_vba(excel_path))
    try:
        sheet = workbook[sheet_name]
        header_row = None
        min_col = None
        max_col = None

        for row_number in range(1, min(sheet.max_row, max_header_row) + 1):
            values = [clean_text(sheet.cell(row_number, column).value) for column in range(1, sheet.max_column + 1)]
            normalized = {normalize_for_compare(value) for value in values if value}
            if {normalize_for_compare(value) for value in required_headers}.issubset(normalized):
                populated_columns = [index + 1 for index, value in enumerate(values) if value]
                header_row = row_number
                min_col = min(populated_columns)
                max_col = max(populated_columns)
                break

        if header_row is None or min_col is None or max_col is None:
            return pd.DataFrame()

        headers = [clean_text(sheet.cell(header_row, column).value) for column in range(min_col, max_col + 1)]
        normalized_row_keys = {
            normalize_for_compare(header)
            for header in ([key_header, *(row_key_headers or [])])
            if clean_text(header)
        }
        key_index = next(
            (
                index
                for index, header in enumerate(headers)
                if normalize_for_compare(header) == normalize_for_compare(key_header)
            ),
            None,
        )
        if key_index is None:
            return pd.DataFrame()
        row_key_indexes = [
            index
            for index, header in enumerate(headers)
            if normalize_for_compare(header) in normalized_row_keys
        ]

        last_row = header_row
        for row_number in range(header_row + 1, sheet.max_row + 1):
            if any(clean_text(sheet.cell(row_number, min_col + index).value) for index in row_key_indexes):
                last_row = row_number

        rows = [
            [sheet.cell(row_number, column).value for column in range(min_col, max_col + 1)]
            for row_number in range(header_row, last_row + 1)
        ]
        return dataframe_from_excel_rows(rows)
    finally:
        workbook.close()


def read_payment_map_items(excel_path: Path, sheet_name: str) -> pd.DataFrame:
    table = read_excel_table(excel_path, sheet_name, "Tabela5")
    if not table.empty:
        return table

    return read_excel_header_region(
        excel_path,
        sheet_name,
        required_headers={"ATO", "PROJETISTA", "VALOR"},
        key_header="PROJETISTA",
    )


def read_measurements_sheet(excel_path: Path, sheet_name: str) -> pd.DataFrame:
    df = read_excel_header_region(
        excel_path,
        sheet_name,
        required_headers={"Número da Medição", "Projeto Referente", "PROJETISTA"},
        key_header="Número da Medição",
        row_key_headers=[
            "PROJETISTA",
            "Motivo Desconto",
            "Descrição do Desconto",
            "DESCRIÇÃO DO DESCONTO",
            "Valor Desconto",
            "Valor do Desconto",
            "VALOR DO DESCONTO",
        ],
    )
    if not df.empty:
        return normalize_measurements_layout(df)

    fallback = pd.read_excel(excel_path, sheet_name=sheet_name, dtype=object).dropna(how="all").reset_index(drop=True)
    return normalize_measurements_layout(fallback)


def read_bm_aux_sheet(excel_path: Path, sheet_name: str) -> pd.DataFrame:
    workbook = load_workbook(excel_path, read_only=False, data_only=True, keep_vba=should_keep_vba(excel_path))
    try:
        if sheet_name not in workbook.sheetnames:
            return pd.DataFrame()
    finally:
        workbook.close()

    table = read_excel_table(excel_path, sheet_name, ["TabelaAuxiliar", "BM_AUX"])
    if not table.empty:
        return normalize_bm_aux_layout(table)

    df = read_excel_header_region(
        excel_path,
        sheet_name,
        required_headers={"Projeto", "Num_Cliente", "Valor De Medição"},
        key_header="Num_Cliente",
    )
    return normalize_bm_aux_layout(df)


def looks_like_document_code(value: Any) -> bool:
    text_value = clean_text(value)
    if not text_value:
        return False
    normalized = normalize_for_compare(text_value)
    return any(marker in normalized for marker in ("grd-", "orc-", "he-", "-e-", "-c-", "-n-"))


def should_apply_positioned_measurements_layout(df: pd.DataFrame) -> bool:
    if df.empty or len(df.columns) < 38:
        return False

    sample = df.head(20)
    coordinator_values = sample.iloc[:, 6].tolist()
    current_professional_values = sample.iloc[:, 28].tolist()
    document_like = sum(1 for value in coordinator_values if looks_like_document_code(value))
    professional_like = sum(1 for value in current_professional_values if clean_text(value))
    return document_like >= 3 and professional_like >= 3


POSITIONED_MEASUREMENTS_HEADERS = [
    "Número da Medição",
    "Projeto Referente",
    "Título Primário",
    "Centro de Custo",
    "Coordenador",
    "Colaborador de Apoio",
    "Número do Documento",
    "Evidência",
    "Data de Cadastro",
    "Formato",
    "Quantidade",
    "Multiplicador",
    "Equivalente (A1 ou Horas)",
    "Porcentagem de Revisão",
    "Emissão Inicial",
    "Retorno Vale",
    "Encerramento",
    "Arquivamento",
    "Medido (Horas)",
    "Item da QQP",
    "Valor Unitário",
    "Valor Bruto",
    "Valor Total",
    "OBS",
    "Valor do Reajuste",
    "Ignorar 1",
    "Localização",
    "Ignorar 2",
    "PROJETISTA",
    "Status Colaborador",
    "% EMISSÃO",
    "CONTRATO",
    "TIPO2",
    "CONDIÇÃO",
    "Valor Auxiliar 1",
    "Valor Auxiliar 2",
    "Percentual Auxiliar 1",
    "VALOR DE MEDIÇÃO",
]


def normalize_measurements_layout(df: pd.DataFrame) -> pd.DataFrame:
    if not should_apply_positioned_measurements_layout(df):
        return df

    normalized_rows = [normalize_positioned_measurement_row(row) for _, row in df.iterrows()]
    return pd.DataFrame(normalized_rows).dropna(how="all").reset_index(drop=True)


def cell_at(row: pd.Series, position: int) -> Any:
    index = position - 1
    if index < 0 or index >= len(row):
        return None
    return row.iloc[index]


def positive_or_fallback(primary: Any, fallback: Any) -> Any:
    primary_value = clean_decimal(primary, default=None)
    if primary_value is not None and primary_value > 0:
        return primary

    fallback_value = clean_decimal(fallback, default=None)
    if fallback_value is not None and fallback_value > 0:
        return fallback

    return primary if primary_value is not None else fallback


def normalize_positioned_measurement_row(row: pd.Series) -> dict[str, Any]:
    raw_values = {f"Origem Coluna {index}": json_safe(value) for index, value in enumerate(row.tolist(), start=1)}

    if looks_like_document_code(cell_at(row, 7)):
        valor_medicao = positive_or_fallback(cell_at(row, 35), cell_at(row, 38))
        return {
            "Número da Medição": cell_at(row, 1),
            "Projeto Referente": cell_at(row, 2),
            "Título Primário": cell_at(row, 3),
            "Centro de Custo": cell_at(row, 4),
            "Localização": cell_at(row, 27),
            "CONTRATO": cell_at(row, 32),
            "Coordenador": cell_at(row, 5),
            "PROJETISTA": cell_at(row, 29),
            "FUNÇÃO": cell_at(row, 8),
            "Mesclado": cell_at(row, 10),
            "Número do Documento": cell_at(row, 7),
            "Evidência": cell_at(row, 8),
            "Data de Cadastro": cell_at(row, 9),
            "Formato": cell_at(row, 10),
            "Quantidade": cell_at(row, 11),
            "Multiplicador": cell_at(row, 12),
            "Equivalente (A1 ou Horas)": cell_at(row, 13),
            "Porcentagem de Revisão": cell_at(row, 14),
            "Emissão Inicial": cell_at(row, 15),
            "Retorno Vale": cell_at(row, 16),
            "Encerramento": cell_at(row, 17),
            "Arquivamento": cell_at(row, 18),
            "Medido (Horas)": cell_at(row, 19),
            "Item da QQP": cell_at(row, 20),
            "Valor Unitário": cell_at(row, 21),
            "Valor Bruto": cell_at(row, 22),
            "Valor Total": valor_medicao,
            "OBS": cell_at(row, 24),
            "Valor do Reajuste": cell_at(row, 25),
            "REFERÊNCIA": cell_at(row, 31),
            "% EMISSÃO": cell_at(row, 31),
            "TIPO2": cell_at(row, 33),
            "CONDIÇÃO": cell_at(row, 34),
            "VALOR DE MEDIÇÃO": valor_medicao,
            **raw_values,
        }

    valor_medicao = positive_or_fallback(cell_at(row, 35), cell_at(row, 38))
    profissional = cell_at(row, 29) or cell_at(row, 7) or cell_at(row, 8)
    return {
        "Número da Medição": cell_at(row, 1),
        "Projeto Referente": cell_at(row, 2),
        "Título Primário": cell_at(row, 3),
        "Centro de Custo": cell_at(row, 4),
        "Localização": cell_at(row, 27),
        "CONTRATO": cell_at(row, 32),
        "Coordenador": cell_at(row, 5),
        "PROJETISTA": profissional,
        "FUNÇÃO": cell_at(row, 26),
        "Mesclado": cell_at(row, 10),
        "Número do Documento": cell_at(row, 7),
        "Evidência": cell_at(row, 8),
        "Data de Cadastro": cell_at(row, 9),
        "Formato": cell_at(row, 10),
        "Quantidade": cell_at(row, 11),
        "Multiplicador": cell_at(row, 12),
        "Equivalente (A1 ou Horas)": cell_at(row, 13),
        "Porcentagem de Revisão": cell_at(row, 14),
        "Emissão Inicial": cell_at(row, 15),
        "Retorno Vale": cell_at(row, 17),
        "Encerramento": None,
        "Arquivamento": None,
        "Medido (Horas)": cell_at(row, 13) or cell_at(row, 11),
        "Item da QQP": cell_at(row, 20),
        "Valor Unitário": cell_at(row, 21),
        "Valor Bruto": cell_at(row, 22),
        "Valor Total": valor_medicao,
        "OBS": cell_at(row, 26),
        "Valor do Reajuste": 0,
        "REFERÊNCIA": cell_at(row, 31),
        "% EMISSÃO": cell_at(row, 31),
        "TIPO2": cell_at(row, 33),
        "CONDIÇÃO": cell_at(row, 34),
        "VALOR DE MEDIÇÃO": valor_medicao,
        **raw_values,
    }


def should_apply_positioned_bm_aux_layout(df: pd.DataFrame) -> bool:
    if df.empty or len(df.columns) < 28:
        return False
    sample = df.head(20)
    first_columns_empty = sample.iloc[:, 0].isna().sum() >= 3 and sample.iloc[:, 1].isna().sum() >= 3
    third_column_has_project = sum(1 for value in sample.iloc[:, 2].tolist() if clean_text(value)) >= 3
    return first_columns_empty and third_column_has_project


POSITIONED_BM_AUX_HEADERS = [
    "Ignorar 1",
    "Ignorar 2",
    "Projeto",
    "Fase",
    "Num_Cliente",
    "Responsavel",
    "Auxiliar",
    "Data_Entrega",
    "Tipo_Revisao",
    "Tipo_Emissao",
    "Data_Emissao",
    "Evidencia_Emissao",
    "Status_Retorno",
    "Formato",
    "Quantidade",
    "Perc_Revisao",
    "Equivalente_Revisado",
    "Tipo_Doc",
    "Contrato",
    "Orçamento",
    "Ordem_Emissao",
    "Ultima_Emissao",
    "Ciclo",
    "CICLO RETORNO",
    "Mesclado",
    "VALOR",
    "% EMISSÃO",
    "VALOR DA MEDIÇÃO",
]


def normalize_bm_aux_layout(df: pd.DataFrame) -> pd.DataFrame:
    if not should_apply_positioned_bm_aux_layout(df):
        return df

    normalized = df.copy()
    normalized = normalized.iloc[:, : len(POSITIONED_BM_AUX_HEADERS)]
    normalized.columns = POSITIONED_BM_AUX_HEADERS[: len(normalized.columns)]
    return normalized


def normalize_cycle(value: Any) -> str | None:
    cleaned = clean_text(value)
    if not cleaned:
        return None
    if cleaned.endswith(".0"):
        cleaned = cleaned[:-2]
    return cleaned


def build_payment_map_item(row: pd.Series, ordem: int, canonical_codes: dict[str, str], ciclo: str = "2605") -> dict[str, Any] | None:
    raw = extract(row, PAYMENT_MAP_ITEM_COLUMNS)
    projetista_codigo = clean_text(raw["projetista_codigo"])
    if not projetista_codigo:
        return None

    normalized_code = normalize_for_compare(projetista_codigo)
    invalid_codes = {"projetista", "total", "valor"}
    if normalized_code in invalid_codes or not any(char.isalpha() for char in projetista_codigo):
        return None

    projetista_codigo = canonical_codes.get(normalized_code, projetista_codigo)
    valor = clean_decimal(raw["valor"])
    ato = clean_text(raw["ato"])
    if has_positive_payment_value(valor) and not ato:
        ato = "PRODUÇÃO"

    payload = {
        "ciclo": ciclo,
        "ordem": ordem,
        "ato": ato,
        "projetista_codigo": projetista_codigo,
        "responsavel": clean_text(raw["responsavel"]),
        "cpf_cnpj": encrypt_sensitive(raw["cpf_cnpj"]),
        "razao_social": clean_text(raw["razao_social"]),
        "intr_sossego": clean_decimal(raw["intr_sossego"]),
        "salobo": clean_decimal(raw["salobo"]),
        "acg": clean_decimal(raw["acg"]),
        "escadas_alumar": clean_decimal(raw["escadas_alumar"]),
        "valor": valor,
        "rev": clean_decimal(raw["rev"]),
        "status": clean_text(raw["status"]),
    }
    raw_payload = safe_raw_payload(row, "unnamed_")
    payload["raw_payload"] = raw_payload
    payload["source_row_hash"] = source_hash({"ciclo": ciclo, "ordem": ordem, **raw_payload})
    return payload


CONTRACT_PAYMENT_COLUMNS = {
    "intr sossego": "intr_sossego",
    "integridade sossego": "intr_sossego",
    "salobo": "salobo",
    "acg": "acg",
    "escadas alumar": "escadas_alumar",
}

CRISTIANO_FIXED_WITH_DOCUMENTS = Decimal("8640")
CRISTIANO_FIXED_WITHOUT_DOCUMENTS = Decimal("12000")

FIXED_CONDITION_REFERENCE = {
    "mauricio spindola": Decimal("8640"),
    "ronald leal": Decimal("21300"),
}

NAME_STOP_WORDS = {"da", "de", "do", "das", "dos", "e", "ltda", "me", "eireli"}


def name_tokens(value: Any) -> list[str]:
    normalized = normalize_for_compare(clean_text(value))
    return [token for token in normalized.split() if len(token) >= 2 and token not in NAME_STOP_WORDS]


def edit_distance(left: str, right: str) -> int:
    previous = list(range(len(right) + 1))
    for i, left_char in enumerate(left, start=1):
        current = [i]
        for j, right_char in enumerate(right, start=1):
            current.append(
                previous[j - 1]
                if left_char == right_char
                else min(previous[j - 1], previous[j], current[j - 1]) + 1
            )
        previous = current
    return previous[-1]


def similar_name_token(left: str, right: str) -> bool:
    if left == right:
        return True
    min_length = min(len(left), len(right))
    if min_length < 4:
        return False
    return edit_distance(left, right) <= (2 if min_length >= 8 else 1)


def same_person_name(left: Any, right: Any) -> bool:
    left_normalized = normalize_for_compare(clean_text(left))
    right_normalized = normalize_for_compare(clean_text(right))
    if not left_normalized or not right_normalized:
        return False
    if left_normalized == right_normalized:
        return True
    if min(len(left_normalized), len(right_normalized)) >= 8 and (
        left_normalized.startswith(right_normalized) or right_normalized.startswith(left_normalized)
    ):
        return True

    left_tokens = name_tokens(left)
    right_tokens = name_tokens(right)
    if not left_tokens or not right_tokens:
        return False
    matches = sum(1 for token in left_tokens if any(similar_name_token(token, right_token) for right_token in right_tokens))
    required = 2 if min(len(left_tokens), len(right_tokens)) >= 2 else 1
    return matches >= required and matches / min(len(left_tokens), len(right_tokens)) >= 0.5


def fixed_condition_reference_for(value: Any) -> Decimal:
    normalized = normalize_for_compare(clean_text(value))
    for key, amount in FIXED_CONDITION_REFERENCE.items():
        if same_person_name(normalized, key):
            return amount
    return Decimal("0")


def is_cristiano_jeferson(value: Any) -> bool:
    return same_person_name(value, "cristiano jeferson")


def fixed_condition_reference_for_payment(item: dict[str, Any]) -> Decimal:
    if is_cristiano_jeferson(item.get("codigo")) or is_cristiano_jeferson(item.get("nome_completo")):
        return CRISTIANO_FIXED_WITH_DOCUMENTS if clean_decimal(item.get("documentos")) > 0 else CRISTIANO_FIXED_WITHOUT_DOCUMENTS
    return fixed_condition_reference_for(item.get("codigo")) or fixed_condition_reference_for(item.get("nome_completo"))


def payment_contract_column(value: Any) -> str | None:
    normalized = normalize_for_compare(clean_text(value))
    if not normalized:
        return None
    if normalized in CONTRACT_PAYMENT_COLUMNS:
        return CONTRACT_PAYMENT_COLUMNS[normalized]
    for key, column in CONTRACT_PAYMENT_COLUMNS.items():
        if key in normalized or normalized in key:
            return column
    return None


def payment_contract_label(column: str | None) -> str:
    return {
        "intr_sossego": "Intr. Sossego",
        "salobo": "Salobo",
        "acg": "ACG",
        "escadas_alumar": "Escadas Alumar",
    }.get(column or "", "PRODUÇÃO")


def latest_cadastros_by_collaborator(conn) -> dict[str, dict[str, Any]]:
    rows = conn.execute(
        text(
            """
            select distinct on (coalesce(colaborador_codigo, responsavel))
                   colaborador_codigo,
                   responsavel,
                   razao_social,
                   cnpj_normalizado,
                   tipo_contrato,
                   valor_condicao_fixa
              from cadastros_fornecedores
             order by coalesce(colaborador_codigo, responsavel), updated_at desc
            """
        )
    ).mappings().all()

    cadastros: dict[str, dict[str, Any]] = {}
    for row in rows:
        keys = [
            normalize_for_compare(row.get("colaborador_codigo")),
            normalize_for_compare(row.get("responsavel")),
        ]
        cadastro = dict(row)
        for key in keys:
            if key and key not in cadastros:
                cadastros[key] = cadastro
    return cadastros


def find_cadastro_for_generated_payment(cadastros: dict[str, dict[str, Any]], item: dict[str, Any]) -> dict[str, Any] | None:
    candidates = [item.get("codigo"), item.get("nome"), item.get("nome_completo")]
    for candidate in candidates:
        direct = cadastros.get(normalize_for_compare(candidate))
        if direct:
            return direct

    for cadastro in cadastros.values():
        for candidate in candidates:
            if (
                same_person_name(candidate, cadastro.get("colaborador_codigo"))
                or same_person_name(candidate, cadastro.get("responsavel"))
                or same_person_name(candidate, cadastro.get("razao_social"))
            ):
                return cadastro
    return None


def generate_payment_map_from_measurements(conn, mapa_pagamento_itens, ciclo: str) -> int:
    rows = conn.execute(
        text(
            """
            select
                coalesce(p.codigo, p.nome) as codigo,
                p.nome as nome,
                p.nome_completo as nome_completo,
                pr.contrato as contrato,
                m.tipo2 as tipo,
                sum(coalesce(m.valor_medicao, 0)) as valor,
                sum(coalesce(m.medido_horas, 0)) as horas
            from medicoes m
            left join profissionais p on p.id = m.id_profissional
            left join projetos pr on pr.id = m.id_projeto
            where m.ciclo = :ciclo
              and m.id_profissional is not null
            group by coalesce(p.codigo, p.nome), p.nome, p.nome_completo, pr.contrato, m.tipo2
            """
        ),
        {"ciclo": ciclo},
    ).mappings().all()

    grouped: dict[str, dict[str, Any]] = {}
    for row in rows:
        codigo = clean_text(row.get("codigo"))
        if not codigo:
            continue
        key = normalize_for_compare(codigo)
        item = grouped.setdefault(
            key,
            {
                "codigo": codigo,
                "nome": clean_text(row.get("nome")),
                "nome_completo": clean_text(row.get("nome_completo")),
                "contratos": {
                    "intr_sossego": Decimal("0"),
                    "salobo": Decimal("0"),
                    "acg": Decimal("0"),
                    "escadas_alumar": Decimal("0"),
                },
                "horas": Decimal("0"),
                "documentos": Decimal("0"),
                "descontos": Decimal("0"),
            },
        )

        valor = clean_decimal(row.get("valor"))
        horas = clean_decimal(row.get("horas"))
        tipo = normalize_for_compare(row.get("tipo"))
        if tipo == "desconto" or valor < 0:
            item["descontos"] += abs(valor)
        else:
            item["documentos"] += valor
            item["horas"] += horas
            column = payment_contract_column(row.get("contrato"))
            if column:
                item["contratos"][column] += valor

    cadastros = latest_cadastros_by_collaborator(conn)
    loaded = 0
    for ordem, item in enumerate(grouped.values(), start=1):
        cadastro = find_cadastro_for_generated_payment(cadastros, item)
        cadastro_valor_fixo = clean_decimal(cadastro.get("valor_condicao_fixa") if cadastro else None)
        referencia_valor_fixo = fixed_condition_reference_for_payment(item)
        is_cristiano = is_cristiano_jeferson(item["codigo"]) or is_cristiano_jeferson(item["nome_completo"])
        valor_fixo = referencia_valor_fixo if is_cristiano else max(cadastro_valor_fixo, referencia_valor_fixo)
        total_documentos = item["documentos"] - item["descontos"]
        valor_total = valor_fixo + total_documentos
        if valor_fixo == 0 and item["documentos"] == 0 and item["descontos"] == 0:
            continue

        contract_values = item["contratos"]
        dominant_column = max(contract_values, key=lambda column: contract_values[column])
        dominant_value = contract_values[dominant_column]
        ato = payment_contract_label(dominant_column if dominant_value > 0 else None)
        base_rateio = item["documentos"] if item["documentos"] > 0 else Decimal("0")
        ratios = {
            column: (contract_values[column] / base_rateio if base_rateio > 0 else Decimal("0"))
            for column in contract_values
        }

        responsavel = clean_text(cadastro.get("responsavel") if cadastro else None) or item["nome_completo"] or item["codigo"]
        razao_social = clean_text(cadastro.get("razao_social") if cadastro else None)
        cnpj = format_cnpj(cadastro.get("cnpj_normalizado") if cadastro else None)
        raw_payload = {
            "origem": "calculado_documentos",
            "documentos": str(item["documentos"]),
            "descontos": str(item["descontos"]),
            "condicoesFixas": {
                "valorFixo": str(valor_fixo),
                "tipoContratacao": clean_text(cadastro.get("tipo_contrato") if cadastro else None) or ("FIXO (PJ)" if valor_fixo > 0 else None),
                "adicionaisFixos": None,
                "observacoesContrato": "Gerado automaticamente por Documentos e Documentos Auxiliares.",
            },
        }

        payment_item = {
            "ciclo": ciclo,
            "ordem": ordem,
            "ato": ato,
            "projetista_codigo": item["codigo"],
            "responsavel": responsavel,
            "cpf_cnpj": encrypt_sensitive(cnpj),
            "razao_social": razao_social,
            "intr_sossego": ratios["intr_sossego"],
            "salobo": ratios["salobo"],
            "acg": ratios["acg"],
            "escadas_alumar": ratios["escadas_alumar"],
            "horas": item["horas"],
            "valor": valor_total,
            "rev": Decimal("0"),
            "status": "PENDENTE",
            "raw_payload": raw_payload,
            "source_row_hash": source_hash({"ciclo": ciclo, "origem": "calculado_documentos", "codigo": item["codigo"]}),
        }
        upsert_payment_map_item(conn, mapa_pagamento_itens, payment_item)
        loaded += 1

    return loaded


def bm_aux_people(raw: dict[str, Any], canonical_codes: dict[str, str]) -> list[tuple[str, str]]:
    people: list[tuple[str, str]] = []
    seen: set[str] = set()
    for role, key in [("RESPONSAVEL", "responsavel"), ("AUXILIAR", "auxiliar")]:
        name = clean_text(raw[key])
        if not name:
            continue
        normalized = normalize_for_compare(name)
        if normalized in seen:
            continue
        seen.add(normalized)
        if BM_AUX_ALLOWED_COLLABORATORS and normalized not in BM_AUX_ALLOWED_COLLABORATORS:
            continue
        people.append((role, canonical_codes.get(normalized, name)))
    return people


def is_bm_aux_only_collaborator(value: Any, canonical_codes: dict[str, str]) -> bool:
    name = clean_text(value)
    if not name:
        return False
    normalized = normalize_for_compare(name)
    canonical = canonical_codes.get(normalized)
    return normalized in BM_AUX_ALLOWED_COLLABORATORS or normalize_for_compare(canonical) in BM_AUX_ALLOWED_COLLABORATORS


def build_bm_aux_measurement(
    row: pd.Series,
    row_number: int,
    codigo: str,
    papel: str,
    ciclo: str,
) -> tuple[dict[str, Any], dict[str, Any]] | None:
    raw = extract(row, BM_AUX_COLUMNS)
    numero_documento = clean_text(raw["numero_cliente"])
    codigo_projeto = clean_text(raw["projeto"]) or clean_text(raw["orcamento"])
    if not numero_documento or not codigo_projeto:
        return None

    valor_medicao = clean_decimal(raw["valor_medicao"])
    valor_unitario = clean_decimal(raw["valor"])
    equivalente = clean_decimal(raw["equivalente_revisado"])
    percentual_emissao = clean_percent(raw["percentual_emissao"])
    raw_payload = safe_raw_payload(row, "unnamed_")
    raw_payload.update({"origem": "BM AUX", "papel": papel, "colaborador": codigo})

    measurement = {
        "numero_medicao": "BM AUX",
        "mesclado": clean_text(raw["mesclado"]),
        "numero_documento": numero_documento,
        "evidencia": clean_text(raw["evidencia_emissao"]),
        "data_cadastro": clean_date(raw["data_emissao"]) or clean_date(raw["data_entrega"]),
        "formato": clean_text(raw["formato"]),
        "quantidade": clean_decimal(raw["quantidade"]),
        "multiplicador": Decimal("1"),
        "equivalente_a1_horas": equivalente,
        "porcentagem_revisao": clean_percent(raw["perc_revisao"]),
        "emissao_inicial": percentual_emissao,
        "retorno_vale": None,
        "encerramento": None,
        "arquivamento": None,
        "medido_horas": equivalente,
        "item_qqp": None,
        "valor_unitario": valor_unitario,
        "valor_bruto": valor_medicao,
        "valor_total": valor_medicao,
        "obs": clean_text(raw["status_retorno"]),
        "valor_reajuste": Decimal("0"),
        "ciclo": ciclo,
        "referencia": clean_text(raw["tipo_revisao"]),
        "percentual_emissao": percentual_emissao,
        "tipo2": clean_text(raw["tipo_doc"]),
        "condicao": clean_text(raw["valor"]),
        "valor_medicao": valor_medicao,
        "raw_payload": raw_payload,
        "source_row_hash": source_hash({"ciclo": ciclo, "origem": "BM AUX", "linha": row_number, "papel": papel, **raw_payload}),
    }

    project = {
        "codigo_projeto": codigo_projeto,
        "titulo_primario": clean_text(raw["orcamento"]),
        "centro_custo": clean_text(raw["fase"]),
        "localizacao": None,
        "contrato": clean_text(raw["contrato"]),
    }

    return measurement, project


def build_bm_aux_record(
    row: pd.Series,
    row_number: int,
    codigo: str,
    papel: str,
    ciclo: str,
) -> dict[str, Any]:
    raw = extract(row, BM_AUX_COLUMNS)
    raw_payload = safe_raw_payload(row, "unnamed_")
    raw_payload.update({"papel": papel, "colaborador": codigo})
    return {
        "responsavel_codigo": codigo,
        "ciclo": ciclo,
        "equivalente_revisado": clean_decimal(raw["equivalente_revisado"]),
        "valor_medicao": clean_decimal(raw["valor_medicao"]),
        "raw_payload": raw_payload,
        "source_row_hash": source_hash({"ciclo": ciclo, "origem": "BM AUX RESUMO", "linha": row_number, "papel": papel, **raw_payload}),
    }


def upsert_bm_aux_medicao(conn, bm_aux_medicoes, data: dict[str, Any]) -> None:
    stmt = insert(bm_aux_medicoes).values(**data)
    stmt = stmt.on_conflict_do_update(
        index_elements=[bm_aux_medicoes.c.source_row_hash],
        set_={
            "responsavel_codigo": stmt.excluded.responsavel_codigo,
            "ciclo": stmt.excluded.ciclo,
            "equivalente_revisado": stmt.excluded.equivalente_revisado,
            "valor_medicao": stmt.excluded.valor_medicao,
            "raw_payload": stmt.excluded.raw_payload,
            "updated_at": text("now()"),
        },
    )
    conn.execute(stmt)


def upsert_payment_map_item(conn, mapa_pagamento_itens, data: dict[str, Any]) -> None:
    stmt = insert(mapa_pagamento_itens).values(**data)
    stmt = stmt.on_conflict_do_update(
        index_elements=[mapa_pagamento_itens.c.source_row_hash],
        set_={
            "ordem": stmt.excluded.ordem,
            "ato": stmt.excluded.ato,
            "projetista_codigo": stmt.excluded.projetista_codigo,
            "responsavel": stmt.excluded.responsavel,
            "cpf_cnpj": stmt.excluded.cpf_cnpj,
            "razao_social": stmt.excluded.razao_social,
            "intr_sossego": stmt.excluded.intr_sossego,
            "salobo": stmt.excluded.salobo,
            "acg": stmt.excluded.acg,
            "escadas_alumar": stmt.excluded.escadas_alumar,
            "horas": stmt.excluded.horas,
            "valor": stmt.excluded.valor,
            "rev": stmt.excluded.rev,
            "status": stmt.excluded.status,
            "raw_payload": stmt.excluded.raw_payload,
            "updated_at": text("now()"),
        },
    )
    conn.execute(stmt)


def upsert_payment_map_context(conn, mapa_pagamento_contexto, data: dict[str, Any]) -> None:
    stmt = insert(mapa_pagamento_contexto).values(**data)
    stmt = stmt.on_conflict_do_update(
        index_elements=[mapa_pagamento_contexto.c.ciclo],
        set_={
            "mes_referencia": stmt.excluded.mes_referencia,
            "producao_label": stmt.excluded.producao_label,
            "producao_inicio": stmt.excluded.producao_inicio,
            "producao_fim": stmt.excluded.producao_fim,
            "ato_label": stmt.excluded.ato_label,
            "ato_ciclo": stmt.excluded.ato_ciclo,
            "contratos": stmt.excluded.contratos,
            "rateio": stmt.excluded.rateio,
            "updated_at": text("now()"),
        },
    )
    conn.execute(stmt)


def build_measurement(row: pd.Series) -> dict[str, Any]:
    raw_measurement = extract(row, MEASUREMENT_COLUMNS)
    payload = {
        "numero_medicao": clean_text(raw_measurement["numero_medicao"]),
        "mesclado": clean_text(raw_measurement["mesclado"]),
        "numero_documento": clean_text(raw_measurement["numero_documento"]),
        "evidencia": clean_text(raw_measurement["evidencia"]),
        "data_cadastro": clean_date(raw_measurement["data_cadastro"]),
        "formato": clean_text(raw_measurement["formato"]),
        "quantidade": clean_decimal(raw_measurement["quantidade"]),
        "multiplicador": clean_decimal(raw_measurement["multiplicador"]),
        "equivalente_a1_horas": clean_decimal(raw_measurement["equivalente_a1_horas"]),
        "porcentagem_revisao": clean_percent(raw_measurement["porcentagem_revisao"]),
        "emissao_inicial": clean_percent(raw_measurement["emissao_inicial"]),
        "retorno_vale": clean_percent(raw_measurement["retorno_vale"]),
        "encerramento": clean_percent(raw_measurement["encerramento"]),
        "arquivamento": clean_percent(raw_measurement["arquivamento"]),
        "medido_horas": clean_decimal(raw_measurement["medido_horas"]),
        "item_qqp": clean_text(raw_measurement["item_qqp"]),
        "valor_unitario": clean_decimal(raw_measurement["valor_unitario"]),
        "valor_bruto": clean_decimal(raw_measurement["valor_bruto"]),
        "valor_total": clean_decimal(raw_measurement["valor_total"]),
        "obs": clean_text(raw_measurement["obs"]),
        "valor_reajuste": clean_decimal(raw_measurement["valor_reajuste"]),
        "ciclo": clean_text(raw_measurement["ciclo"]),
        "referencia": clean_text(raw_measurement["referencia"]),
        "percentual_emissao": clean_percent(raw_measurement["percentual_emissao"]),
        "tipo2": clean_text(raw_measurement["tipo2"]),
        "condicao": clean_text(raw_measurement["condicao"]),
        "valor_medicao": clean_decimal(raw_measurement["valor_medicao"]),
    }
    raw_payload = safe_raw_payload(row, "Unnamed")
    payload["raw_payload"] = raw_payload
    payload["source_row_hash"] = source_hash(raw_payload)
    return payload


def build_discount_measurement(row: pd.Series, base_measurement: dict[str, Any], ciclo: str) -> dict[str, Any] | None:
    raw_measurement = extract(row, MEASUREMENT_COLUMNS)
    valor_desconto = clean_decimal(raw_measurement["valor_desconto"], default=None)
    if valor_desconto is None or valor_desconto == 0:
        return None

    desconto = abs(valor_desconto)
    motivo = clean_text(raw_measurement["motivo_desconto"]) or clean_text(base_measurement.get("obs")) or "Desconto"
    payload = {
        "numero_medicao": base_measurement.get("numero_medicao"),
        "mesclado": base_measurement.get("mesclado"),
        "numero_documento": "DESCONTO",
        "evidencia": base_measurement.get("evidencia"),
        "data_cadastro": base_measurement.get("data_cadastro"),
        "formato": None,
        "quantidade": Decimal("1"),
        "multiplicador": Decimal("1"),
        "equivalente_a1_horas": Decimal("1"),
        "porcentagem_revisao": None,
        "emissao_inicial": None,
        "retorno_vale": None,
        "encerramento": None,
        "arquivamento": None,
        "medido_horas": Decimal("0"),
        "item_qqp": None,
        "valor_unitario": desconto,
        "valor_bruto": desconto,
        "valor_total": desconto,
        "obs": motivo,
        "valor_reajuste": Decimal("0"),
        "ciclo": ciclo,
        "referencia": base_measurement.get("referencia"),
        "percentual_emissao": Decimal("1"),
        "tipo2": "DESCONTO",
        "condicao": str(desconto),
        "valor_medicao": desconto,
    }
    raw_payload = safe_raw_payload(row, "Unnamed")
    raw_payload["__linha_gerada"] = "DESCONTO"
    raw_payload["Motivo Desconto"] = motivo
    raw_payload["Valor Desconto"] = str(desconto)
    payload["raw_payload"] = raw_payload
    payload["source_row_hash"] = source_hash({"ciclo": ciclo, **raw_payload})
    return payload


def has_discount_data(row: pd.Series) -> bool:
    raw_measurement = extract(row, MEASUREMENT_COLUMNS)
    valor_desconto = clean_decimal(raw_measurement["valor_desconto"], default=None)
    motivo = clean_text(raw_measurement["motivo_desconto"])
    return bool((valor_desconto is not None and valor_desconto != 0) or motivo)


def build_discount_only_base_measurement(row: pd.Series, ciclo: str) -> dict[str, Any]:
    raw_measurement = extract(row, MEASUREMENT_COLUMNS)
    raw_payload = safe_raw_payload(row, "Unnamed")
    payload = {
        "numero_medicao": clean_text(raw_measurement["numero_medicao"]) or "DESCONTO",
        "mesclado": clean_text(raw_measurement["mesclado"]),
        "numero_documento": "DESCONTO",
        "evidencia": clean_text(raw_measurement["evidencia"]),
        "data_cadastro": clean_date(raw_measurement["data_cadastro"]),
        "formato": None,
        "quantidade": Decimal("1"),
        "multiplicador": Decimal("1"),
        "equivalente_a1_horas": Decimal("1"),
        "porcentagem_revisao": None,
        "emissao_inicial": None,
        "retorno_vale": None,
        "encerramento": None,
        "arquivamento": None,
        "medido_horas": Decimal("0"),
        "item_qqp": None,
        "valor_unitario": Decimal("0"),
        "valor_bruto": Decimal("0"),
        "valor_total": Decimal("0"),
        "obs": clean_text(raw_measurement["motivo_desconto"]),
        "valor_reajuste": Decimal("0"),
        "ciclo": ciclo,
        "referencia": clean_text(raw_measurement["referencia"]),
        "percentual_emissao": Decimal("1"),
        "tipo2": "DESCONTO",
        "condicao": "0",
        "valor_medicao": Decimal("0"),
        "raw_payload": raw_payload,
        "source_row_hash": source_hash({"ciclo": ciclo, **raw_payload}),
    }
    return payload


def ingest(
    excel_path: Path,
    sheet_name: str,
    base_sheet_name: str,
    payment_map_sheet_name: str,
    bm_aux_sheet_name: str,
    database_url: str,
    create_schema: bool,
    full_refresh: bool,
    ciclo: str | None = None,
) -> dict[str, int]:
    engine = create_engine(database_url, future=True)
    if create_schema:
        load_schema(engine, Path(__file__).resolve().parents[1] / "database" / "schema.sql")

    sheet_name = resolve_sheet_name(excel_path, sheet_name, SHEET_ALIASES["Documentos"])
    base_sheet_name = resolve_optional_sheet_name(excel_path, base_sheet_name, SHEET_ALIASES["Base"])
    payment_map_sheet_name = resolve_optional_sheet_name(
        excel_path,
        payment_map_sheet_name,
        SHEET_ALIASES["MAPA PAGTO"],
    )
    bm_aux_sheet_name = resolve_sheet_name(
        excel_path,
        bm_aux_sheet_name,
        SHEET_ALIASES["Documentos Auxiliares"],
    )

    df = read_measurements_sheet(excel_path, sheet_name)
    base_df = (
        pd.read_excel(excel_path, sheet_name=base_sheet_name, dtype=object).dropna(how="all").reset_index(drop=True)
        if base_sheet_name
        else pd.DataFrame()
    )
    payment_map_items_df = read_payment_map_items(excel_path, payment_map_sheet_name) if payment_map_sheet_name else pd.DataFrame()
    payment_map_df = payment_map_items_df
    bm_aux_df = read_bm_aux_sheet(excel_path, bm_aux_sheet_name)
    positive_payment_codes = build_positive_payment_codes(df)
    canonical_codes = build_canonical_professional_codes(base_df)
    if not payment_map_items_df.empty:
        canonical_codes.update(build_payment_map_canonical_codes(payment_map_items_df))

    projetos, profissionais, medicoes, mapa_pagamento_contexto, mapa_pagamento_itens, bm_aux_medicoes = reflect_tables(engine)
    payment_context = (
        build_payment_map_context(excel_path, payment_map_sheet_name, ciclo=ciclo)
        if payment_map_sheet_name
        else build_generated_payment_context(df, bm_aux_df, ciclo=ciclo)
    )
    ciclo_efetivo = payment_context["ciclo"]
    affected_collaborator_codes = collect_import_collaborator_codes(df, bm_aux_df, canonical_codes, ciclo_efetivo)
    base_loaded = 0
    payment_status_loaded = 0
    payment_items_loaded = 0
    bm_aux_rows_loaded = 0
    bm_aux_measurements_loaded = 0
    inserted_or_updated = 0
    source_hashes: set[str] = set()
    duplicate_source_rows = 0
    skipped = 0

    with engine.begin() as conn:
        if full_refresh:
            clear_imported_collaborators(conn, ciclo_efetivo, affected_collaborator_codes)

        upsert_payment_map_context(conn, mapa_pagamento_contexto, payment_context)

        for _, row in base_df.iterrows():
            base_professional = build_base_professional(row)
            if not base_professional:
                continue
            upsert_base_professional(conn, profissionais, base_professional)
            base_loaded += 1

        conn.execute(text("update profissionais set status_colaborador = null"))
        if not payment_map_df.empty:
            for _, row in payment_map_df.iterrows():
                payment_status = build_payment_map_status(row, positive_payment_codes, canonical_codes)
                if not payment_status:
                    continue
                upsert_payment_map_status(conn, profissionais, payment_status)
                payment_status_loaded += 1

            for index, row in payment_map_items_df.iterrows():
                payment_item = build_payment_map_item(row, index + 1, canonical_codes, ciclo=ciclo_efetivo)
                if not payment_item:
                    continue
                upsert_payment_map_item(conn, mapa_pagamento_itens, payment_item)
                payment_items_loaded += 1

        for index, row in bm_aux_df.iterrows():
            raw = extract(row, BM_AUX_COLUMNS)
            row_cycle = normalize_cycle(raw["ciclo"])
            if row_cycle and row_cycle != ciclo_efetivo:
                continue

            people = bm_aux_people(raw, canonical_codes)
            if not people:
                continue
            bm_aux_rows_loaded += 1

            for papel, codigo in people:
                built = build_bm_aux_measurement(row, index + 1, codigo, papel, ciclo_efetivo)
                if not built:
                    continue

                measurement, project = built
                id_projeto = upsert_project(conn, projetos, project)
                id_profissional = upsert_professional(
                    conn,
                    profissionais,
                    {"nome": codigo, "codigo": codigo, "funcao": None},
                )
                measurement.update(
                    {
                        "id_projeto": id_projeto,
                        "id_coordenador": None,
                        "id_profissional": id_profissional,
                    }
                )

                if measurement["source_row_hash"] in source_hashes:
                    duplicate_source_rows += 1
                source_hashes.add(measurement["source_row_hash"])

                stmt = insert(medicoes).values(**measurement)
                stmt = stmt.on_conflict_do_update(
                    index_elements=[medicoes.c.source_row_hash],
                    set_={
                        column: stmt.excluded[column]
                        for column in measurement
                        if column not in {"source_row_hash"}
                    }
                    | {"updated_at": text("now()")},
                )
                conn.execute(stmt)
                inserted_or_updated += 1
                bm_aux_measurements_loaded += 1

                bm_aux_record = build_bm_aux_record(row, index + 1, codigo, papel, ciclo_efetivo)
                upsert_bm_aux_medicao(conn, bm_aux_medicoes, bm_aux_record)

        for _, row in df.iterrows():
            project_raw = extract(row, PROJECT_COLUMNS)
            codigo_projeto = clean_text(project_raw["codigo_projeto"])
            numero_medicao = clean_text(first_value(row, MEASUREMENT_COLUMNS["numero_medicao"]))
            valid_measurement = is_valid_measurement_key(numero_medicao, codigo_projeto)
            discount_only = not valid_measurement and has_discount_data(row)
            if not valid_measurement and not discount_only:
                skipped += 1
                continue
            professional_raw = extract(row, PROFESSIONAL_COLUMNS)
            if is_bm_aux_only_collaborator(professional_raw["nome"], canonical_codes) and not discount_only:
                skipped += 1
                continue

            contrato = clean_text(project_raw["contrato"])
            if discount_only:
                codigo_projeto = f"DESCONTO-{normalize_for_compare(contrato).upper() or 'GERAL'}"

            id_projeto = upsert_project(
                conn,
                projetos,
                {
                    "codigo_projeto": codigo_projeto,
                    "titulo_primario": clean_text(project_raw["titulo_primario"]),
                    "centro_custo": clean_text(project_raw["centro_custo"]),
                    "localizacao": clean_text(project_raw["localizacao"]),
                    "contrato": contrato,
                },
            )

            id_profissional = upsert_professional(conn, profissionais, professional_raw)
            id_coordenador = upsert_professional(conn, profissionais, extract(row, COORDINATOR_COLUMNS))
            if discount_only:
                measurement = build_discount_only_base_measurement(row, ciclo_efetivo)
            else:
                measurement = build_measurement(row)
                # O ciclo efetivo da carga é a fonte da verdade. Isso evita manter
                # linhas antigas quando a planilha é atualizada com outro ciclo.
                measurement["ciclo"] = ciclo_efetivo
                measurement["source_row_hash"] = source_hash(
                    {"ciclo": ciclo_efetivo, **measurement["raw_payload"]}
                )
                measurement.update(
                    {
                        "id_projeto": id_projeto,
                        "id_coordenador": id_coordenador,
                        "id_profissional": id_profissional,
                    }
                )
                if measurement["source_row_hash"] in source_hashes:
                    duplicate_source_rows += 1
                source_hashes.add(measurement["source_row_hash"])

                stmt = insert(medicoes).values(**measurement)
                stmt = (
                    stmt
                    .on_conflict_do_update(
                        index_elements=[medicoes.c.source_row_hash],
                        set_={
                            column: stmt.excluded[column]
                            for column in measurement
                            if column not in {"source_row_hash"}
                        }
                        | {"updated_at": text("now()")},
                    )
                )
                conn.execute(stmt)
                inserted_or_updated += 1

            discount_measurement = build_discount_measurement(row, measurement, ciclo_efetivo)
            if discount_measurement:
                discount_measurement.update(
                    {
                        "id_projeto": id_projeto,
                        "id_coordenador": id_coordenador,
                        "id_profissional": id_profissional,
                    }
                )
                if discount_measurement["source_row_hash"] in source_hashes:
                    duplicate_source_rows += 1
                source_hashes.add(discount_measurement["source_row_hash"])

                discount_stmt = insert(medicoes).values(**discount_measurement)
                discount_stmt = (
                    discount_stmt
                    .on_conflict_do_update(
                        index_elements=[medicoes.c.source_row_hash],
                        set_={
                            column: discount_stmt.excluded[column]
                            for column in discount_measurement
                            if column not in {"source_row_hash"}
                        }
                        | {"updated_at": text("now()")},
                    )
                )
                conn.execute(discount_stmt)
                inserted_or_updated += 1

        if payment_map_df.empty:
            payment_items_loaded = generate_payment_map_from_measurements(conn, mapa_pagamento_itens, ciclo_efetivo)
            payment_status_loaded = payment_items_loaded

    return {
        "rows_read": len(df),
        "base_sheet_found": 1 if base_sheet_name else 0,
        "base_rows_read": len(base_df),
        "base_professionals_loaded": base_loaded,
        "payment_map_rows_read": len(payment_map_df),
        "payment_status_loaded": payment_status_loaded,
        "payment_map_items_rows_read": len(payment_map_items_df),
        "payment_map_items_loaded": payment_items_loaded,
        "bm_aux_rows_read": len(bm_aux_df),
        "bm_aux_rows_loaded": bm_aux_rows_loaded,
        "bm_aux_measurements_loaded": bm_aux_measurements_loaded,
        "rows_processed": inserted_or_updated,
        "rows_unique_by_source_hash": len(source_hashes),
        "rows_duplicate_by_source_hash": duplicate_source_rows,
        "rows_skipped": skipped,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="ETL relacional de medições da aba Documentos.")
    parser.add_argument("--excel", default=DEFAULT_EXCEL_PATH, help="Caminho do arquivo .xlsm/.xlsx.")
    parser.add_argument("--sheet", default=DEFAULT_SHEET, help="Nome da aba de origem.")
    parser.add_argument("--base-sheet", default=DEFAULT_BASE_SHEET, help="Nome da aba opcional de cadastro de profissionais.")
    parser.add_argument("--payment-map-sheet", default=DEFAULT_PAYMENT_MAP_SHEET, help="Nome da aba do mapa de pagamento.")
    parser.add_argument("--bm-aux-sheet", default=DEFAULT_BM_AUX_SHEET, help="Nome da aba auxiliar de BM.")
    parser.add_argument(
        "--append",
        action="store_true",
        help="Não limpa o banco antes da carga. Use apenas para cargas incrementais controladas.",
    )
    parser.add_argument(
        "--ciclo",
        default=None,
        help="Ciclo no formato YYMM (ex: 2606). Se omitido, deriva automaticamente do campo 'Mês referência' da planilha.",
    )
    parser.add_argument(
        "--database-url",
        default=os.getenv("ETL_DATABASE_URL") or os.getenv("DATABASE_URL"),
        help="URL PostgreSQL. Ex: postgresql+psycopg2://usuario:senha@localhost:5432/medicoes",
    )
    parser.add_argument("--create-schema", action="store_true", help="Executa database/schema.sql antes da carga.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if not args.database_url:
        raise SystemExit("Informe --database-url ou defina a variável ETL_DATABASE_URL.")

    result = ingest(
        Path(args.excel),
        args.sheet,
        args.base_sheet,
        args.payment_map_sheet,
        args.bm_aux_sheet,
        args.database_url,
        args.create_schema,
        not args.append,
        ciclo=args.ciclo,
    )
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
