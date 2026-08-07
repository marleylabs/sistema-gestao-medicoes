from __future__ import annotations

import tempfile
from pathlib import Path

from openpyxl import Workbook

from ingest_medicoes import (
    build_generated_payment_context,
    read_bm_aux_sheet,
    read_measurements_sheet,
)


DOCUMENTOS_HEADERS = [
    "Número da Medição",
    "Projeto Referente",
    "Título Primário",
    "Centro de Custo",
    "Coordenador",
    "Líder Responsável",
    "Número do Documento",
    "Evidência",
    "Data de Cadastro",
    "Formato",
    "Quantidade",
    "Multiplicador",
    "Equivalente (A1 ou Horas)",
    "Porcentagem de Revisão",
    "Emissão Inicial",
    "Retorno Vale",
    "Arquivamento",
    "Medido (Horas)",
    "Valor do Reajuste 3º",
    "Item da QQP",
    "Valor Unitário",
    "Valor Bruto",
    "Valor do Reajuste",
    "Valor Total",
    "OBS",
    "FUNÇÃO",
    "Localização",
    "Motivo Desconto",
    "Valor Desconto",
    "CICLO",
    "PROJETISTA",
    "REFERÊNCIA",
    "% EMISSÃO",
    "CONTRATO",
    "TIPO",
    "CONDIÇÃO",
    "VALOR DE MEDIÇÃO",
]

DOCUMENTOS_AUXILIARES_HEADERS = [
    "Projeto",
    "Fase",
    "Num_Cliente",
    "Responsavel",
    "Auxiliar",
    "Data_Entrega",
    "Tipo_Revisao",
    "Tipo_Emissao",
    "Data_Emissao",
    "Evidencia_Emissao",
    "Status_Retorno",
    "Formato",
    "Quantidade",
    "Perc_Revisao",
    "Equivalente_Revisado",
    "Tipo_Doc",
    "Contrato",
    "Orçamento",
    "Ordem_Emissao",
    "Ultima_Emissao",
    "Ciclo",
    "Ciclo Retorno",
    "Mesclado",
    "Valor",
    "% Emissão",
    "Valor De Medição",
]


def create_mask(path: Path) -> None:
    workbook = Workbook()
    instruction = workbook.active
    instruction.title = "Instruções"
    instruction.append(["Painel", "Orientação"])

    documents = workbook.create_sheet("Documentos")
    documents.append(DOCUMENTOS_HEADERS)

    aux = workbook.create_sheet("Documentos Auxiliares")
    aux.append(DOCUMENTOS_AUXILIARES_HEADERS)

    workbook.save(path)
    workbook.close()


def assert_columns(name: str, actual: list[str], expected: list[str]) -> None:
    if actual != expected:
        raise AssertionError(f"{name}: colunas divergentes.\nEsperado: {expected}\nAtual: {actual}")


def main() -> None:
    with tempfile.TemporaryDirectory() as temp_dir:
        excel_path = Path(temp_dir) / "Mascara_Importacao_Medicoes.xlsx"
        create_mask(excel_path)

        documents = read_measurements_sheet(excel_path, "Documentos")
        aux = read_bm_aux_sheet(excel_path, "Documentos Auxiliares")
        context = build_generated_payment_context(documents, aux, ciclo="2607")

    assert documents.shape == (0, len(DOCUMENTOS_HEADERS))
    assert aux.shape == (0, len(DOCUMENTOS_AUXILIARES_HEADERS))
    assert context["ciclo"] == "2607"
    assert context["contratos"] == []
    assert context["rateio"] == []

    assert_columns("Documentos", list(documents.columns), DOCUMENTOS_HEADERS)
    assert_columns("Documentos Auxiliares", list(aux.columns), DOCUMENTOS_AUXILIARES_HEADERS)

    print("OK: padrão de importação sem MAPA PAGTO validado.")


if __name__ == "__main__":
    main()
